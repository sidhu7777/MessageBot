"""
REQ-006: Doctor Appointment List — Excel Sheet Generation & Sending
Verifies that:
  1. _build_doctor_report_xlsx() generates a valid .xlsx file
  2. The file has correct headers (Booking Number, Patient Name, Contact, etc.)
  3. All patient rows are written correctly
  4. The scheduler calls send_document_fn with the xlsx path
  5. The dedup key prevents sending the same schedule reminder twice

Run: python tests/req_006_doctor_excel_list.py
"""
import os
import sys
import tempfile
import threading
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

PASS = 0
FAIL = 0


def check(label: str, condition: bool, detail: str = "") -> None:
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  [PASS] {label}")
    else:
        FAIL += 1
        detail_str = f" -- {detail}" if detail else ""
        print(f"  [FAIL] {label}{detail_str}")


# ─── Fake DoctorReminder rows ──────────────────────────────────────────────────

@dataclass
class FakeDoctorReminder:
    appointment_id: int
    doctor_whatsapp: str
    doctor_telegram_chat_id: str
    patient_name: str
    patient_contact: str
    clinic_name: str
    slot_date: str
    slot_time: str
    status: str
    booking_number: Optional[int]
    schedule_id: int
    schedule_start_time: str
    schedule_end_time: str


def make_rows(n: int = 5) -> list[FakeDoctorReminder]:
    rows = []
    for i in range(1, n + 1):
        rows.append(FakeDoctorReminder(
            appointment_id=100 + i,
            doctor_whatsapp="whatsapp:+919392569600",
            doctor_telegram_chat_id="8299824956",
            patient_name=f"Patient {i}",
            patient_contact=f"987654321{i}",
            clinic_name="City Care Clinic",
            slot_date="2026-03-01",
            slot_time="10:00",
            status="BOOKED",
            booking_number=i,
            schedule_id=1,
            schedule_start_time="10:00",
            schedule_end_time="12:00",
        ))
    return rows


# ─── Test 1: Excel file is generated ──────────────────────────────────────────

def test_excel_file_generated():
    print("\n[TEST] Excel file is created with correct path")

    from src.automation.scheduler import AutomationScheduler

    scheduler = AutomationScheduler(
        booking_repository=MagicMock(),
        send_message_fn=lambda *a: None,
        enabled=True,
    )

    rows = make_rows(3)

    with tempfile.TemporaryDirectory() as tmpdir:
        # Patch os.path.join and makedirs to use temp dir
        orig_makedirs = os.makedirs
        orig_join = os.path.join

        def fake_join(*parts):
            if parts[0] == "data":
                return orig_join(tmpdir, *parts[1:])
            return orig_join(*parts)

        with patch("src.automation.scheduler.os.makedirs", lambda p, exist_ok=False: orig_makedirs(p, exist_ok=True)):
            with patch("src.automation.scheduler.os.path.join", side_effect=fake_join):
                path = scheduler._build_doctor_report_xlsx(
                    rows=rows,
                    to_number="telegram:8299824956",
                    slot_date="2026-03-01",
                    schedule_id=1,
                    start_time="10:00",
                    end_time="12:00",
                )

        check("xlsx path returned", path.endswith(".xlsx"), f"path={path}")

        # File must exist in temp dir or original data/reports
        xlsx_exists = os.path.exists(path) or os.path.exists(
            os.path.join(ROOT, path) if not os.path.isabs(path) else path
        )
        # Try relative from ROOT
        full_path = path if os.path.isabs(path) else os.path.join(ROOT, path)
        check("xlsx file created on disk", os.path.exists(full_path), f"expected={full_path}")


# ─── Test 2: Excel file has correct headers ───────────────────────────────────

def test_excel_correct_headers():
    print("\n[TEST] Excel file has all required column headers")

    import openpyxl
    from src.automation.scheduler import AutomationScheduler

    scheduler = AutomationScheduler(
        booking_repository=MagicMock(),
        send_message_fn=lambda *a: None,
        enabled=True,
    )

    os.makedirs(os.path.join(ROOT, "data", "reports"), exist_ok=True)

    path = scheduler._build_doctor_report_xlsx(
        rows=make_rows(2),
        to_number="test_headers",
        slot_date="2026-03-01",
        schedule_id=999,
        start_time="10:00",
        end_time="11:00",
    )

    full_path = path if os.path.isabs(path) else os.path.join(ROOT, path)
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

    required_headers = [
        "Booking Number",
        "Patient Name",
        "Contact",
        "Clinic",
        "Appointment Date",
        "Appointment Time",
        "Status",
    ]

    for h in required_headers:
        check(f"header '{h}' present", h in headers, f"headers found: {headers}")


# ─── Test 3: All patient rows are written ─────────────────────────────────────

def test_excel_patient_rows_written():
    print("\n[TEST] All patient rows written to Excel correctly")

    import openpyxl
    from src.automation.scheduler import AutomationScheduler

    scheduler = AutomationScheduler(
        booking_repository=MagicMock(),
        send_message_fn=lambda *a: None,
        enabled=True,
    )

    n_patients = 5
    rows = make_rows(n_patients)

    os.makedirs(os.path.join(ROOT, "data", "reports"), exist_ok=True)

    path = scheduler._build_doctor_report_xlsx(
        rows=rows,
        to_number="test_rows",
        slot_date="2026-03-01",
        schedule_id=998,
        start_time="10:00",
        end_time="11:00",
    )
    full_path = path if os.path.isabs(path) else os.path.join(ROOT, path)

    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    data_rows = ws.max_row - 1  # subtract header row

    check(f"all {n_patients} patients written (got {data_rows})", data_rows == n_patients)

    # Check first data row has correct patient name
    first_name = ws.cell(row=2, column=2).value
    check("first patient name is 'Patient 1'", first_name == "Patient 1", f"got: {first_name}")

    # Check booking numbers
    for i in range(1, n_patients + 1):
        booking_num = ws.cell(row=1 + i, column=1).value
        check(f"row {i} booking number is {i}", booking_num == i, f"got: {booking_num}")


# ─── Test 4: send_document_fn called with xlsx path ───────────────────────────

def test_send_document_fn_called():
    print("\n[TEST] scheduler calls send_document_fn with xlsx path when due")

    from src.automation.scheduler import AutomationScheduler

    sent_docs = []

    def send_doc_fn(to, path, caption):
        sent_docs.append({"to": to, "path": path, "caption": caption})

    mock_booking_repo = MagicMock()

    # Make reminder due right now
    now = datetime.now()
    lead = 10  # minutes
    # Schedule window 10 minutes from now (center of window)
    fake_start = (now + timedelta(minutes=lead)).strftime("%H:%M")
    fake_date = now.strftime("%Y-%m-%d")

    rows = make_rows(3)
    for r in rows:
        r.slot_date = fake_date
        r.schedule_start_time = fake_start
        r.schedule_end_time = (now + timedelta(minutes=lead + 30)).strftime("%H:%M")

    mock_booking_repo.list_due_doctor_reminders.return_value = rows

    os.makedirs(os.path.join(ROOT, "data", "reports"), exist_ok=True)

    scheduler = AutomationScheduler(
        booking_repository=mock_booking_repo,
        send_message_fn=lambda *a: None,
        send_document_fn=send_doc_fn,
        enabled=True,
        doctor_reminder_enabled=True,
        doctor_reminder_lead_minutes=lead,
        doctor_reminder_window_seconds=600,  # wide window so timing doesn't flake
    )

    # Patch the key store to always say "not seen" so it sends
    with patch.object(scheduler._reminder_keys, "has", return_value=False):
        with patch.object(scheduler._reminder_keys, "add"):
            scheduler._run_reminders_once()

    check("send_document_fn was called", len(sent_docs) >= 1,
          f"sent_docs={sent_docs}")
    if sent_docs:
        check("sent path ends in .xlsx", sent_docs[0]["path"].endswith(".xlsx"),
              f"path={sent_docs[0]['path']}")
        check("sent to correct destination",
              sent_docs[0]["to"] in {"telegram:8299824956", "whatsapp:+919392569600"},
              f"to={sent_docs[0]['to']}")


# ─── Test 5: Dedup key prevents second send ───────────────────────────────────

def test_dedup_prevents_second_reminder():
    print("\n[TEST] Dedup key prevents sending same schedule reminder twice")

    from src.automation.scheduler import AutomationScheduler

    call_count = [0]

    def send_doc_fn(to, path, caption):
        call_count[0] += 1

    mock_booking_repo = MagicMock()
    now = datetime.now()
    lead = 10
    fake_start = (now + timedelta(minutes=lead)).strftime("%H:%M")
    fake_date = now.strftime("%Y-%m-%d")

    rows = make_rows(2)
    for r in rows:
        r.slot_date = fake_date
        r.schedule_start_time = fake_start
        r.schedule_end_time = (now + timedelta(minutes=lead + 30)).strftime("%H:%M")

    mock_booking_repo.list_due_doctor_reminders.return_value = rows

    scheduler = AutomationScheduler(
        booking_repository=mock_booking_repo,
        send_message_fn=lambda *a: None,
        send_document_fn=send_doc_fn,
        enabled=True,
        doctor_reminder_lead_minutes=lead,
        doctor_reminder_window_seconds=600,
    )

    # First run: simulate key NOT seen → should send
    with patch.object(scheduler._reminder_keys, "has", return_value=False):
        with patch.object(scheduler._reminder_keys, "add"):
            scheduler._run_reminders_once()

    first_run_calls = call_count[0]
    check("first run sends reminder", first_run_calls >= 1)

    # Second run: key IS already seen → should NOT send again
    with patch.object(scheduler._reminder_keys, "has", return_value=True):
        scheduler._run_reminders_once()

    check("second run does NOT resend (dedup)", call_count[0] == first_run_calls,
          f"calls after second run={call_count[0]}, expected={first_run_calls}")


# ─── Test 6: Dr. Sanjay — 48 patients, 1PM-5PM, T-10 = 12:50 PM ──────────────

def test_dr_sanjay_t10_both_platforms():
    """
    Real-world scenario:
      Doctor : Dr. Sanjay (doctor_id=1)
      Slot   : today 13:00 – 17:00  (4 hours, 5-min intervals → 48 appointments)
      T-10   : datetime.now() frozen at 12:50 PM  →  delta = exactly 600 s
      Send   : both telegram:8299824956  AND  whatsapp:+919392569600
    """
    print("\n[TEST] Dr. Sanjay T-10: 48 patients (1PM-5PM, 5-min slots) — both Telegram & WhatsApp")

    import openpyxl
    from datetime import datetime as real_datetime
    from unittest.mock import patch, MagicMock
    from src.automation.scheduler import AutomationScheduler

    TODAY = real_datetime.now().strftime("%Y-%m-%d")   # 2026-02-26
    SLOT_START = "13:00"
    SLOT_END   = "17:00"   # 1PM to 5PM = 4 hours × 12 slots/hr = 48 slots
    DR_TG      = "8299824956"
    DR_WA      = "whatsapp:+919392569600"
    FAKE_NOW   = real_datetime.strptime(f"{TODAY} 12:50", "%Y-%m-%d %H:%M")

    # Build 48 rows: 5-min intervals from 13:00 to 15:55
    slot_times = []
    h, m = 13, 0
    while (h, m) < (17, 0):
        slot_times.append(f"{h:02d}:{m:02d}")
        m += 5
        if m >= 60:
            m -= 60
            h += 1
    assert len(slot_times) == 48, f"Expected 48 slots, got {len(slot_times)}"

    rows = []
    for i, slot_t in enumerate(slot_times, start=1):
        rows.append(FakeDoctorReminder(
            appointment_id=200 + i,
            doctor_whatsapp=DR_WA,
            doctor_telegram_chat_id=DR_TG,
            patient_name=f"Patient {i:02d}",
            patient_contact=f"9{i:09d}",
            clinic_name="City Care Clinic",
            slot_date=TODAY,
            slot_time=slot_t,
            status="BOOKED",
            booking_number=i,
            schedule_id=1,
            schedule_start_time=SLOT_START,
            schedule_end_time=SLOT_END,
        ))

    check("48 fake appointment rows created", len(rows) == 48, f"got {len(rows)}")

    sent_docs: list[dict] = []

    def send_doc_fn(to, path, caption):
        sent_docs.append({"to": to, "path": path, "caption": caption})

    mock_repo = MagicMock()
    mock_repo.list_due_doctor_reminders.return_value = rows

    os.makedirs(os.path.join(ROOT, "data", "reports"), exist_ok=True)

    scheduler = AutomationScheduler(
        booking_repository=mock_repo,
        send_message_fn=lambda *a: None,
        send_document_fn=send_doc_fn,
        enabled=True,
        doctor_reminder_enabled=True,
        doctor_reminder_lead_minutes=10,
        doctor_reminder_window_seconds=60,   # ±60 s window around T-10
    )

    # Freeze datetime.now() to 12:50 PM — exactly T-10 before 13:00
    class _FakeDatetime(real_datetime):
        @classmethod
        def now(cls, tz=None):
            return FAKE_NOW

    with patch("src.automation.scheduler.datetime", _FakeDatetime):
        with patch.object(scheduler._reminder_keys, "has", return_value=False):
            with patch.object(scheduler._reminder_keys, "add"):
                scheduler._run_reminders_once()

    # ── Delivery checks ────────────────────────────────────────────────────────
    destinations_sent_to = {d["to"] for d in sent_docs}

    check("send_document_fn called at least twice (Telegram + WhatsApp)",
          len(sent_docs) >= 2, f"calls={len(sent_docs)}")
    check("Telegram destination received the list",
          f"telegram:{DR_TG}" in destinations_sent_to,
          f"destinations={destinations_sent_to}")
    check("WhatsApp destination received the list",
          DR_WA in destinations_sent_to,
          f"destinations={destinations_sent_to}")

    # ── Excel content check ────────────────────────────────────────────────────
    if sent_docs:
        xlsx_path = sent_docs[0]["path"]
        full = xlsx_path if os.path.isabs(xlsx_path) else os.path.join(ROOT, xlsx_path)
        check("report file exists on disk", os.path.exists(full), f"path={full}")
        if os.path.exists(full):
            wb = openpyxl.load_workbook(full)
            ws = wb.active
            data_rows = ws.max_row - 1   # minus header
            check("Excel contains all 48 patient rows", data_rows == 48,
                  f"got {data_rows} rows")
            # Spot-check first slot time = 13:00
            first_time = ws.cell(row=2, column=6).value   # column 6 = Appointment Time
            check("first appointment time is 01:00 PM", first_time in ("13:00", "01:00 PM"),
                  f"got {first_time!r}")
            last_time = ws.cell(row=49, column=6).value
            check("last appointment time is 04:55 PM", last_time in ("16:55", "04:55 PM"),
                  f"got {last_time!r}")

    # ── Caption check ─────────────────────────────────────────────────────────
    if sent_docs:
        caption = sent_docs[0]["caption"]
        check("caption mentions 10 minutes", "10" in caption,
              f"caption={caption}")
        check("caption mentions total patients (48)", "48" in caption,
              f"caption={caption}")


if __name__ == "__main__":
    print("=" * 60)
    print("REQ-006: Doctor Appointment List — Excel Generation & Send")
    print("=" * 60)

    test_excel_file_generated()
    test_excel_correct_headers()
    test_excel_patient_rows_written()
    test_send_document_fn_called()
    test_dedup_prevents_second_reminder()
    test_dr_sanjay_t10_both_platforms()

    print("\n" + "=" * 60)
    print(f"RESULT: {PASS} passed, {FAIL} failed out of {PASS + FAIL} checks")
    print("=" * 60)
    sys.exit(0 if FAIL == 0 else 1)
